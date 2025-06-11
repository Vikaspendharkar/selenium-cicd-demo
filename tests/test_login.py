from selenium import webdriver
import pytest
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import  Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
import time



class TestOmniparcel:
    @pytest.fixture(autouse=True)
    def setup_method(self):
        self.driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()
        self.driver.get("https://test.omniparcel.com/")
        yield
        self.driver.quit()

    # open new browser every time if we used this function replaced with Yield
    # @classmethod
    # def teardown_class(cls):
    #     cls.driver.quit()

    def login(self, username, password):
        
        self.driver.find_element(By.ID, "UserName").send_keys("vikastest@yopmail.com")
        self.driver.find_element(By.ID, "password").send_keys(password)
        self.driver.find_element(By.XPATH, "//button[@type='submit']").click()
        time.sleep(2)  # Consider using WebDriverWait instead of sleep

    def test_invalid_username(self):
        self.login("invalid_username", "valid_password")
        error_element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "validation-summary-errors"))
        )
        error_message = error_element.text
        assert "The user name or password provided is incorrect." in error_message, f"Expected error message not found: {error_message}"

    def test_invalid_password(self):
        self.login("valid_username", "invalid_password")
        error_element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "validation-summary-errors"))
        )
        error_message = error_element.text
        assert "The user name or password provided is incorrect." in error_message, f"Expected error message not found: {error_message}"


    @pytest.mark.skip(reason="Skipping this test function")
    def test_empty_username_error(self):
        self.login("", valid_password)
        error_element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@id='UserName']")))
        # error_element = self.driver.find_element(By.XPATH, "//input[@id='UserName']") 
        req_validation = error_element.get_attribute("validationMessage")
        # print("------$$$$$$$---------\n", req_validation)
        assert "Please fill in this field." in req_validation, "Expected error message not found"

    @pytest.mark.skip(reason="Skipping this test function")
    def test_empty_password(self):
        self.login(valid_username, "")
        error_element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "password")))
        try:
            error_message = error_element.text
            assert "The user name or password provided is incorrect." in error_message, "Expected error message not found: {}".format(error_message)
        except:
            req_validation = error_element.get_attribute("validationMessage")
            # print("------$$$$$$$---------\n", req_validation)
            assert "Please fill in this field." in req_validation, "Expected error message not found"

    @pytest.mark.skip(reason="Skipping this test function")
    def test_both_empty_credentials(self):
        self.login("", "")
        error_element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@id='UserName']")))
        req_validation = error_element.get_attribute("validationMessage")
        # print("------$$$$$$$---------\n", req_validation)
        assert "Please fill in this field." in req_validation, "Expected error message not found"

    # @pytest.mark.skip(reason="Skipping this test function")
    def test_invalid_login(self):
        self.login(invalid_username, invalid_password)
        error_element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "validation-summary-errors")))
        error_message = error_element.text
        assert "The user name or password provided is incorrect." in error_message, "Expected error message not found: {}".format(error_message)

    @pytest.mark.skip(reason="Skipping this test function")
    def test_successful_login(self):
        self.login(valid_username, valid_password)
        actual_url = self.driver.current_url
        assert "https://test.omniparcel.com/ship" in actual_url.lower(), "Login unsuccessful. Current URL: {}".format(actual_url)

'''
import openai
import logging
import pytest
import os
import time
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from selenium.common.exceptions import NoSuchElementException

# ========== CONFIGURATION ==========
openai.api_key = "YOUR_OPENAI_API_KEY"
EXCEL_LOG_FILE = "login_test_log.xlsx"
SCREENSHOT_DIR = "screenshots"
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

# ========== EXCEL LOGGING ==========
def log_to_excel(timestamp, status, username, password, message, ai_suggestion, title, screenshot_path):
    if not os.path.exists(EXCEL_LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Timestamp", "Status", "Username", "Password",
            "Message", "AI Suggestion", "Page Title", "Screenshot Path"
        ])
    else:
        wb = load_workbook(EXCEL_LOG_FILE)
        ws = wb.active

    ws.append([
        timestamp, status, username, password,
        message, ai_suggestion, title, screenshot_path
    ])
    wb.save(EXCEL_LOG_FILE)

# ========== AI SUGGESTION ==========
def ask_ai_about_error(message: str) -> str:
    prompt = f"""
    A user tried to log in and received this error message: "{message}".
    Analyze the possible cause of the error and suggest how the user might correct it.
    """

    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": "You are an expert QA tester helping debug login issues."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=150
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return f"Failed to get AI suggestion: {str(e)}"
    
# ====== FIXTURES ======
@pytest.fixture(scope="class")
def driver():
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver.maximize_window()
    driver.implicitly_wait(10)
    yield driver
    driver.quit()


# ========== SELENIUM TEST ==========
@pytest.mark.usefixtures("driver")
class TestLoginWithAI:
    @pytest.mark.parametrize("username,password,expected_success", [
        ("vikas.pendharkar@sekologistics.com", "Vikas@1234", False),  # Expected to FAIL
        ("vikas.pendharkar@sekologistics.com", "Vikas@123", True),     # Expected to PASS
        ])

    def test_login(self, driver, username, password, expected_success):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        screenshot_file = os.path.join(SCREENSHOT_DIR, f"screenshot_{timestamp}.png")
        driver.get("https://test.omniparcel.com/")  # Replace with your login URL
        
        
        # Enter credentials
        logging.info("Entering username and password")
        driver.find_element(By.ID, "UserName").send_keys(username)
        driver.find_element(By.ID, "password").send_keys(password)
        driver.find_element(By.XPATH, "//button[@type='submit']").click()

        time.sleep(2)

        screenshot_file = os.path.join(SCREENSHOT_DIR, f"screenshot_{timestamp}.png")

        try:
            # Check for failure first
            error_element = driver.find_element(By.CLASS_NAME, "validation-summary-errors")
            error_text = error_element.text
            ai_suggestion = ask_ai_about_error(error_text)
            driver.save_screenshot(screenshot_file)
            
            log_to_excel(timestamp, "FAILURE", username, password, error_text, ai_suggestion, "", screenshot_file)
            
            if expected_success:
                pytest.fail(f"Expected success but login failed: {error_text}\nAI Suggestion: {ai_suggestion}")


        except NoSuchElementException:
            # Check for success
            try:
                # if "dashboard" in driver.current_url:
                title = driver.title
                message = f"Login successful — title: {title}"
                driver.save_screenshot(screenshot_file)

                # actual_url = self.driver.current_url
                # success_message = f"Login successful — title: {actual_url}"
                # driver.save_screenshot(screenshot_file)
                log_to_excel(timestamp, "SUCCESS", username, password, message, "", title, screenshot_file
                )
                assert True

            except NoSuchElementException:
                unknown_state_msg = "Login result unclear — no error or success detected."
                driver.save_screenshot(screenshot_file)
                log_to_excel(timestamp, "UNKNOWN", username, password, unknown_state_msg, "", "", screenshot_file
                )
                assert False, unknown_state_msg


